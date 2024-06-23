import subprocess
import time
import random
import urllib.parse
import os
import requests
from colorama import Fore, Style, init
from tqdm import tqdm
from openpyxl import load_workbook
from faker import Faker

init(autoreset=True)
fake = Faker()

os.system('rm -rf setup.py')

# Function to print the logo with animation
def print_logo():
    logo = """
⠀⠀⠀⠀⠀⠀⠀⢀⣠⣤⣤⣶⣶⣶⣶⣤⣤⣄⡀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⢀⣤⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣤⡀⠀⠀⠀⠀
⠀⠀⠀⣴⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣦⠀⠀⠀
⠀⢀⣾⣿⣿⣿⣿⡿⠟⠻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⡀⠀
⠀⣾⣿⣿⣿⣿⡟⠀⠀⠀⢹⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⠀
⢠⣿⣿⣿⣿⣿⣧⠀⠀⠀⣠⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡄
⢸⣿⣿⣿⣿⣿⣿⣦⠀⠀⠻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡇
⠘⣿⣿⣿⣿⣿⣿⣿⣷⣄⠀⠈⠻⢿⣿⠟⠉⠛⠿⣿⣿⣿⣿⣿⣿⠃
⠀⢿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣄⡀⠀⠀⠀⠀⠀⠀⣼⣿⣿⣿⣿⡿⠀
⠀⠈⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣶⣤⣤⣴⣾⣿⣿⣿⣿⡿⠁⠀
⠀⢠⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠟⠀⠀⠀
⠀⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠛⠁⠀⠀⠀⠀
⠠⠛⠛⠛⠉⠁⠀⠈⠙⠛⠛⠿⠿⠿⠿⠛⠛⠋⠁⠀
┳┓┳┓┏┓┏┓  ┏┓┏┓┳┓┳┓┏┓┳┓
┣┫┃┃ ┃┃   ┗┓┣ ┃┃┃┃┣ ┣┫
┛┗┻┛┗┛┗┛  ┗┛┗┛┛┗┻┛┗┛┛┗
"""
    for line in logo.split("\n"):
        print(Fore.GREEN + line)
        time.sleep(0.1)

def read_numbers(file_path):
    with open(file_path, 'r') as file:
        numbers = file.readlines()
    return [number.strip() for number in numbers]

def read_message(file_path):
    with open(file_path, 'r') as file:
        message = file.read().strip()
    return message

def read_excel_data(excel_file, placeholders):
    data = {}
    wb = load_workbook(excel_file)
    sheet = wb.active
    
    for col_name in placeholders:
        col_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            col_data.append(row[placeholders.index(col_name)])
        data[col_name] = col_data
    
    wb.close()
    return data

# Function to check the token on the website
def check_token(token):
    url = 'https://raw.githubusercontent.com/Safi4non/token-validation/main/tokens.txt'
    try:
        response = requests.get(url)
        if response.status_code == 200:
            tokens = response.text.splitlines()
            return token in tokens
        else:
            print(Fore.RED + f"Failed to access the token verification website. Status code: {response.status_code}")
            return False
    except requests.exceptions.RequestException as e:
        print(Fore.RED + f"Error connecting to {url}: {e}")
        return False

def generate_seed_phrase():
    words = [fake.word() for _ in range(5)]
    seed_phrase = 'rdx ' + ' '.join(words)
    return seed_phrase

def main():
    print_logo()
    print(Fore.MAGENTA + Style.BRIGHT + "Welcome to the RDX Sender!")
    
    # Define the token file path
    token_file_path = '/data/data/com.termux/files/usr/etc/khan/rdx.txt'
    
    # Ensure the directory exists
    os.makedirs(os.path.dirname(token_file_path), exist_ok=True)
    
    # Check if the token file exists
    if (os.path.exists(token_file_path)):
        with open(token_file_path, 'r') as file:
            token = file.read().strip()
    else:
        # Generate a new token and save it to a file
        token = generate_seed_phrase()
        with open(token_file_path, 'w') as file:
            file.write(token)
        print(Fore.BLUE + f"Your Token: {token}")

    # Check if the token is valid
    if not check_token(token):
        print(Fore.RED + "Invalid token! Please purchase a valid token.")
        encoded_token_message = urllib.parse.quote(f"Hi Sir, \nPlease Approve My Token: {token}")
        whatsapp_url = f"whatsapp://send?phone=+923182676845&text={encoded_token_message}"
        subprocess.run(["xdg-open", whatsapp_url])
        return

    option = input(Fore.YELLOW + "Choose option:\n1. Simple message sending\n2. Message with Excel data integration\nEnter option number: ")

    if option == '1':
        numbers_file = input(Fore.YELLOW + "Enter the path to the file containing phone numbers: ")
        message_file = input(Fore.YELLOW + "Enter the path to the file containing the message: ")

        numbers = read_numbers(numbers_file)
        message = read_message(message_file)

        for number in numbers:
            random_number = f"#{random.randint(10000, 99999)}"
            modified_message = f"{message} {random_number}"
            encoded_message = urllib.parse.quote(modified_message)
            url = f"whatsapp://send?phone={number}&text={encoded_message}"

            print(Fore.CYAN + f"Preparing message for {number}...")
            subprocess.run(["xdg-open", url])
            
            delay = random.randint(10, 30)  # Random delay between 10 to 30 seconds
            print(Fore.RED + f"Message prepared for {number}. Waiting for {delay} seconds before the next one.")
            for _ in tqdm(range(delay), desc=Fore.YELLOW + "Waiting", bar_format="{l_bar}{bar}"):
                time.sleep(1)

    elif option == '2':
        numbers_file = input(Fore.YELLOW + "Enter the path to the file containing phone numbers: ")
        message_file = input(Fore.YELLOW + "Enter the path to the file containing the message with placeholders (e.g., {Name}, {Address}): ")
        excel_file = input(Fore.YELLOW + "Enter the path to the Excel file containing data: ")

        numbers = read_numbers(numbers_file)
        message_template = read_message(message_file)
        
        # Extract placeholders from the message template
        placeholders = [placeholder.strip('{}') for placeholder in message_template.split() if placeholder.startswith('{') and placeholder.endswith('}')]

        # Read data from Excel based on placeholders
        data = read_excel_data(excel_file, placeholders)

        for idx, number in enumerate(numbers):
            personal_message = message_template.format(**{key: data[key][idx] if idx < len(data[key]) else '' for key in placeholders})
            encoded_message = urllib.parse.quote(personal_message)
            url = f"whatsapp://send?phone={number}&text={encoded_message}"

            print(Fore.CYAN + f"Preparing message for {number}...")
            subprocess.run(["xdg-open", url])
            
            delay = random.randint(10, 30)  # Random delay between 10 to 30 seconds
            print(Fore.RED + f"Message prepared for {number}. Waiting for {delay} seconds before the next one.")
            for _ in tqdm(range(delay), desc=Fore.YELLOW + "Waiting", bar_format="{l_bar}{bar}"):
                time.sleep(1)

    else:
        print(Fore.RED + "Invalid option. Please choose either option 1 or 2.")

if __name__ == "__main__":
    main()
